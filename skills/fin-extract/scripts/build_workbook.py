#!/usr/bin/env python3
"""Build a financial-model-style Excel workbook from batch-extracted FIBRA CSVs.

Layout: periods as columns (B+), line items as rows. One tab per statement.

Usage:
    python build_workbook.py <tables_parent_dir> [--output <path.xlsx>]
"""

import csv
import re
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, numbers
from openpyxl.utils import get_column_letter

QUARTER_RE = re.compile(r"(\d)[QqTt](\d{2})")

# Periods to skip per statement type (corrupted or non-tabular data)
SKIP_PERIODS = {
    "noi_ebitda": {"1Q21", "2Q21"},
    "ffo_affo": {"1Q21", "2Q21"},
}

TAB_CONFIG = [
    ("income_statement", "Income Statement"),
    ("balance_sheet", "Balance Sheet"),
    ("cash_flow", "Cash Flow"),
    ("noi_ebitda", "NOI & EBITDA"),
    ("ffo_affo", "FFO & AFFO"),
]


def period_sort_key(period):
    m = QUARTER_RE.match(period.upper())
    if m:
        return (int(m.group(2)), int(m.group(1)))
    return (99, 99)


def read_csv(csv_path):
    with open(csv_path, "r", encoding="utf-8") as f:
        return list(csv.reader(f))


def clean_value(raw):
    """Parse a dollar-formatted string into a float, or return None."""
    if not raw or not raw.strip():
        return None
    s = raw.strip()
    # Handle parenthetical negatives: ($123,456) or (123,456)
    neg = False
    if s.startswith("(") and s.endswith(")"):
        neg = True
        s = s[1:-1].strip()
    if s.startswith("($") and not neg:
        neg = True
        s = s[2:].rstrip(")")
    s = s.lstrip("$").strip()
    if not s or s == "-" or s == "—":
        return 0.0
    s = s.replace(",", "")
    # Reject strings that don't look like numbers
    if not re.match(r"^-?\d[\d.]*%?$", s):
        return None
    if s.endswith("%"):
        return s  # preserve percentage as string
    try:
        val = float(s)
        return -val if neg else val
    except ValueError:
        return None


def normalize_label(label):
    """Normalize a label for fuzzy matching across quarters."""
    s = label.lower().strip()
    # Strip trailing annotations like (1), (2), (Net), etc.
    s = re.sub(r"\s*\(\d+\)\s*$", "", s)
    s = re.sub(r"\s*\(net\)\s*$", "", s, flags=re.IGNORECASE)
    # Normalize common spelling variants
    s = s.replace("non-operating", "non operating")
    s = s.replace("non-operation", "non operating")
    s = s.replace("non operating", "non operating")
    s = s.replace("non operation", "non operating")
    # Remove trailing periods and colons for comparison
    s = s.rstrip(".:;")
    # Collapse whitespace
    s = re.sub(r"\s+", " ", s)
    # Remove leading dollar signs, bullets, "A: " / "L: " prefixes
    s = s.lstrip("$•·- ")
    # Remove parenthetical qualifiers for matching
    s = re.sub(r"\s*\((own|consolidated)\s*portfolio\)", "", s)
    return s


# ---------------------------------------------------------------------------
# Parsers — each returns [(label, value), ...] for the current period
# ---------------------------------------------------------------------------

def find_header_row(rows):
    """Find the row index containing period patterns like '3Q25' or 'Thousand'."""
    for i, row in enumerate(rows):
        joined = " ".join(row)
        if QUARTER_RE.search(joined):
            return i
        if "thousand" in joined.lower() and ("pesos" in joined.lower() or any(QUARTER_RE.search(c) for c in row)):
            return i
    return None


def find_period_column(header_row, period):
    """Find which column in the header row matches the given period string.

    Handles both Q (English) and T (Spanish "Trimestre") period labels,
    e.g. "2Q25" matches "2T25" and vice versa.
    """
    period_upper = period.upper()
    # Also try the T-variant (Spanish reports use e.g. "2T25" instead of "2Q25")
    t_variant = period_upper.replace("Q", "T")
    for i, cell in enumerate(header_row):
        normalized = cell.upper().replace(" ", "")
        if period_upper in normalized or t_variant in normalized:
            return i
    return None


def parse_standard_csv(rows, period):
    """Parse income statement, NOI/EBITDA, and FFO/AFFO CSVs.

    These have a header row with period labels, label column 0 (or sometimes 1
    for FFO/AFFO with an empty col 0), and values in the period's column.
    """
    header_idx = find_header_row(rows)
    if header_idx is None:
        return []

    header = rows[header_idx]
    period_col = find_period_column(header, period)
    if period_col is None:
        return []

    # Determine label column: usually 0, but some FFO/AFFO CSVs have labels
    # shifted if col 0 is a section header repeated on every row
    label_col = 0

    results = []
    for row in rows[header_idx + 1:]:
        if len(row) <= max(label_col, period_col):
            continue
        label = row[label_col].strip()
        if not label:
            # Try col 1 if col 0 is empty
            if len(row) > 1 and row[1].strip():
                label = row[1].strip()
            else:
                continue

        # Skip metadata/noise rows
        lower_label = label.lower()
        if lower_label.startswith(("note:", "fibra soma", "mexican pesos",
                                   "soma21", "closing price", "outstanding cbfi",
                                   "market capitalization", "mxn$", "interest affo",
                                   "\u2022", "we ", "showing ", "affo ", "to ")):
            continue
        if "days of operation" in lower_label:
            continue
        # Skip narrative text (full sentences that leaked from PDF)
        if len(label) > 60 and " " in label:
            continue
        # Skip orphan fragments (just a year, or very short non-label strings)
        if re.match(r"^\d{4}\.$", label) or re.match(r"^to \d{4}\.$", lower_label):
            continue

        val_str = row[period_col] if period_col < len(row) else ""
        val = clean_value(val_str)

        # Skip rows where value is inline text (corrupted extractions)
        if val is None and val_str.strip():
            # If the "value" cell has lots of text, it's probably narrative
            if len(val_str.strip()) > 30:
                continue

        results.append((label, val))

    return results


def _parse_split_value(row, val_col):
    """Handle values split across two columns (1Q21 full-peso format).
    E.g. col1='$  9' col2='5,827,542' → combine into '$95827542' → 95827.542 (÷1000).
    """
    raw = row[val_col].strip() if val_col < len(row) else ""
    next_col = val_col + 1

    # If the next column has a pure numeric continuation, combine them
    if next_col < len(row):
        c2 = row[next_col].strip().lstrip(",")
        if c2 and re.match(r"^[\d,]+$", c2):
            num_part = raw.lstrip("$").strip().rstrip(",")
            neg = "(" in raw
            num_part = num_part.lstrip("(").rstrip(")")
            combined = (num_part + c2).replace(",", "")
            try:
                val = float(combined)
                if neg:
                    val = -val
                return round(val / 1000, 0)
            except ValueError:
                pass

    val = clean_value(raw)
    # 1Q21 reports values in full pesos
    if val is not None and isinstance(val, float) and abs(val) > 500000:
        val = round(val / 1000, 0)
    return val


EXPENSE_KEYWORDS = frozenset([
    "operation and maintenance", "operation fee", "administrator expenses",
    "property tax", "property insurance", "cost of land sales",
    "total expenses", "other non-operation",
])


def _normalize_expense_signs(pairs):
    """Negate expense line items that were extracted as positive.

    2021 reports present expenses as positive numbers; 2022+ use parenthetical
    negatives. This normalizes everything to negative for expenses.
    """
    # Find the zone between "Total revenues" and "Total expenses"
    in_expense_zone = False
    result = []
    for label, val in pairs:
        lower = label.lower()
        if "total revenues" in lower or "total revenue" in lower:
            in_expense_zone = True
            result.append((label, val))
            continue
        if "total expenses" in lower:
            in_expense_zone = False
            if val is not None and isinstance(val, (int, float)) and val > 0:
                val = -val
            result.append((label, val))
            continue

        if in_expense_zone and val is not None and isinstance(val, (int, float)) and val > 0:
            if any(kw in lower for kw in EXPENSE_KEYWORDS) or in_expense_zone:
                val = -val

        result.append((label, val))
    return result


def parse_income_statement(csv_path, period):
    rows = read_csv(csv_path)

    # 1Q21 has a unique format: full pesos, split values, different headers
    if period.upper() == "1Q21":
        pairs = _parse_income_statement_1q21(rows)
    else:
        pairs = parse_standard_csv(rows, period)

    return _normalize_expense_signs(pairs)


def _parse_income_statement_1q21(rows):
    """Parse 1Q21 income statement: full pesos, values split across cols 1-2."""
    results = []
    # Skip header rows (first 3 rows are metadata)
    start = 0
    for i, row in enumerate(rows):
        if row and row[0].strip().lower().startswith(("base rent", "rental", "tenant")):
            start = i
            break
    if start == 0:
        start = 3

    for row in rows[start:]:
        if not row or not row[0].strip():
            continue
        label = row[0].strip()
        lower = label.lower()
        if lower.startswith(("financial", "mexican", "consolidated")):
            continue

        val = _parse_split_value(row, 1)
        results.append((label, val))

    return results


def parse_noi_ebitda(csv_path, period):
    rows = read_csv(csv_path)
    return parse_standard_csv(rows, period)


def parse_ffo_affo(csv_path, period):
    """Parse FFO/AFFO — uses standard parser but the label column may be 0 or
    offset. Also some CSVs have an empty col between label and values."""
    rows = read_csv(csv_path)
    header_idx = find_header_row(rows)
    if header_idx is None:
        return []

    header = rows[header_idx]
    period_col = find_period_column(header, period)
    if period_col is None:
        return []

    # Detect label column: scan data rows to see if col 0 has diverse labels
    label_col = 0
    col0_vals = set()
    for row in rows[header_idx + 1: header_idx + 10]:
        if row:
            col0_vals.add(row[0].strip())
    # If col 0 is always the same value (section header), labels are in col 1
    if len(col0_vals) <= 2 and any(len(v) > 0 for v in col0_vals):
        non_empty = [v for v in col0_vals if v]
        if len(non_empty) == 1:
            label_col = 1 if len(rows[header_idx + 1]) > 1 else 0

    # Stop at the BIVA section (market data, not part of FFO/AFFO)
    results = []
    for row in rows[header_idx + 1:]:
        if len(row) <= label_col:
            continue
        label = row[label_col].strip()
        if not label:
            if label_col == 0 and len(row) > 1 and row[1].strip():
                label = row[1].strip()
            else:
                continue

        lower = label.lower()
        if "biva" in lower or "closing price" in lower or "outstanding cbfi" in lower:
            break
        if "market capitalization" in lower:
            break
        if lower.startswith(("note:", "fibra soma", "mexican pesos", "soma21")):
            break
        if "days of operation" in lower:
            continue

        val_str = row[period_col] if period_col < len(row) else ""
        val = clean_value(val_str)

        # Skip narrative text masquerading as values
        if val is None and val_str.strip() and len(val_str.strip()) > 30:
            continue

        results.append((label, val))

    return results


def parse_balance_sheet_3q25_plus(rows, period):
    """3Q25+ balance sheet: compact 3-column format.
    Col 0 = label (sometimes combined with value for assets),
    Col 1 = liabilities label or value, Col 2 = liabilities value.
    """
    results = []
    for row in rows:
        if not row or not any(c.strip() for c in row):
            continue
        # Assets side: label in col 0, value embedded in col 0 or col 1
        label0 = row[0].strip() if len(row) > 0 else ""
        cell1 = row[1].strip() if len(row) > 1 else ""
        cell2 = row[2].strip() if len(row) > 2 else ""

        # Skip header rows
        lower0 = label0.lower()
        if any(x in lower0 for x in ["balance sheet", "thousand", "assets as of"]):
            continue

        # Asset lines: value is sometimes stuck in the same cell with the
        # liabilities label. Pattern: "$4,560,740 Current portion..."
        # Or value is cleanly in a separate position.

        # Check if cell1 has a value at the start
        val_match = re.match(r"^(\$[\d,]+(?:\.\d+)?)\s*(.*)", cell1)
        if val_match and label0 and ":" not in label0:
            asset_val = clean_value(val_match.group(1))
            liab_label = val_match.group(2).strip()
            if asset_val is not None:
                results.append((f"A: {label0}", asset_val))
            if liab_label:
                liab_val = clean_value(cell2)
                results.append((f"L: {liab_label}", liab_val))
            continue

        # Pure asset line (no liabilities data in this row)
        if label0 and cell1 and not cell2:
            av = clean_value(cell1)
            if av is not None and ":" not in label0:
                results.append((f"A: {label0}", av))
                continue

        # Liabilities/equity lines: these appear on the right side
        # In 3Q25 format, when left side says "Total non-current assets"
        # repeatedly, the labels are in col 1 and values in col 2
        if cell1 and cell2:
            lv = clean_value(cell2)
            if lv is not None:
                results.append((f"L: {cell1}", lv))
                continue

        # Section headers (end with ":")
        if label0 and label0.endswith(":"):
            results.append((f"A: {label0}", None))
        if cell1 and cell1.endswith(":"):
            results.append((f"L: {cell1}", None))

    return results


def parse_balance_sheet_pre3q25(rows, period):
    """Pre-3Q25 balance sheet: side-by-side layout.
    Typical structure (1Q22-2Q25): 6 columns.
    Col 0 = section group, col 1 = asset label, col 2 = asset value,
    col 3 = liabilities section, col 4 = liab label, col 5 = liab value.
    """
    results = []
    ncols = max((len(r) for r in rows), default=0)

    if ncols <= 6:
        # Standard 6-column layout (3Q21 through ~2Q25)
        for row in rows:
            if len(row) < 3:
                continue
            # Skip header rows
            joined = " ".join(row).lower()
            if "thousand" in joined or "assets" == row[0].strip().lower() or "as of" in joined:
                continue

            # Asset side: label in col 0 or 1, value in col 1 or 2
            a_label = row[1].strip() if len(row) > 1 else ""
            a_val_str = row[2].strip() if len(row) > 2 else ""
            section_label = row[0].strip()

            if not a_label and section_label and ":" not in section_label:
                a_label = section_label
                # Value might be in col 1 or col 2
                a_val_str = row[1].strip() if row[1].strip() else (row[2].strip() if len(row) > 2 else "")

            if a_label:
                av = clean_value(a_val_str)
                if av is not None:
                    results.append((f"A: {a_label}", av))
                elif ":" in a_label:
                    results.append((f"A: {a_label}", None))

            # Liabilities side: col 3/4 = label, col 5 = value
            l_label = ""
            l_val_str = ""
            if len(row) > 5:
                l_label = row[4].strip() if row[4].strip() else row[3].strip()
                l_val_str = row[5].strip()
            elif len(row) > 4:
                l_label = row[3].strip()
                l_val_str = row[4].strip()

            if l_label:
                lv = clean_value(l_val_str)
                if lv is not None:
                    results.append((f"L: {l_label}", lv))
                elif ":" in l_label or l_label.lower().startswith(("trustors", "contributed", "earned")):
                    results.append((f"L: {l_label}", None))

    else:
        # Wide format (2Q24 style: 30+ columns with lots of empty cells)
        # Values are scattered across columns. Find them by looking for
        # dollar amounts.
        for row in rows:
            joined = " ".join(row).lower()
            if "thousand" in joined or not any(c.strip() for c in row):
                continue

            # Find all non-empty cells and their positions
            cells = [(i, c.strip()) for i, c in enumerate(row) if c.strip()]
            if not cells:
                continue

            # Heuristic: first half of columns = assets, second half = liabilities
            mid = ncols // 2
            asset_cells = [(i, c) for i, c in cells if i < mid]
            liab_cells = [(i, c) for i, c in cells if i >= mid]

            # Extract asset label + value
            a_labels = [c for _, c in asset_cells if clean_value(c) is None and c.lower() not in ("assets", "")]
            a_values = [c for _, c in asset_cells if clean_value(c) is not None]
            if a_labels and a_values:
                results.append((f"A: {a_labels[-1]}", clean_value(a_values[-1])))
            elif a_labels and not a_values:
                lbl = a_labels[-1]
                if "as of" not in lbl.lower() and lbl.lower() != "assets":
                    results.append((f"A: {lbl}", None))

            # Extract liabilities label + value
            l_labels = [c for _, c in liab_cells if clean_value(c) is None and c.lower() not in ("", "liabilities")]
            l_values = [c for _, c in liab_cells if clean_value(c) is not None]
            if l_labels and l_values:
                results.append((f"L: {l_labels[-1]}", clean_value(l_values[-1])))
            elif l_labels and not l_values:
                lbl = l_labels[-1]
                if "as of" not in lbl.lower():
                    results.append((f"L: {lbl}", None))

    return results


def parse_balance_sheet_1q21(rows, period):
    """1Q21 balance sheet: full pesos (not thousands), labels in col 0,
    values split across cols 1-2 (formatting artifact)."""
    results = []
    for row in rows:
        if len(row) < 2:
            continue
        label = row[0].strip()
        if not label:
            continue
        lower = label.lower()
        if any(x in lower for x in ["fibra soma", "mexican pesos", "as of"]):
            continue

        # Value may be split: col1 = "$                  3" col2 = "39,014,303"
        # or col1 = "$  614,410,704" col2 = ""
        raw_val = row[1].strip()
        if len(row) > 2 and row[2].strip():
            # Check if col2 is a continuation of the number
            c2 = row[2].strip().lstrip(",")
            if re.match(r"^[\d,]+$", c2):
                # Combine: strip $ and whitespace from col1, append col2
                num_part = raw_val.lstrip("$").strip().rstrip(",").lstrip("(").rstrip(")")
                combined = num_part + c2
                combined = combined.replace(",", "")
                try:
                    val = float(combined)
                    # Check for parenthetical negative
                    if "(" in raw_val:
                        val = -val
                    # Convert full pesos to thousands
                    val = round(val / 1000, 0)
                    # Determine side
                    if "liabilit" in lower or "equity" in lower or "payable" in lower or "debt" in lower or "deposit" in lower or "tenant" in lower or "trust" in lower or "controlling" in lower or "contribution" in lower or "non current liab" in lower or "non-current liab" in lower or "stockholder" in lower:
                        results.append((f"L: {label}", val))
                    else:
                        results.append((f"A: {label}", val))
                    continue
                except ValueError:
                    pass

        val = clean_value(raw_val)
        if val is not None:
            # 1Q21 is in full pesos, convert to thousands
            if isinstance(val, float) and abs(val) > 100000:
                val = round(val / 1000, 0)

            lower_l = label.lower()
            if any(x in lower_l for x in ["liabilit", "equity", "payable", "debt", "deposit",
                                           "tenant admission", "acquisition real estate",
                                           "other accounts pay", "securit", "trust", "contribut",
                                           "controlling", "stockholder", "ipo expenses",
                                           "initial adjustment", "non controling"]):
                results.append((f"L: {label}", val))
            else:
                results.append((f"A: {label}", val))
        elif ":" in label or label.lower().startswith(("current", "non-current", "non current", "stockholder")):
            side = "L" if any(x in label.lower() for x in ["liabilit", "equity", "stockholder"]) else "A"
            results.append((f"{side}: {label}", None))

    return results


def parse_balance_sheet(csv_path, period):
    rows = read_csv(csv_path)
    if not rows:
        return []

    ncols = max(len(r) for r in rows)

    # 1Q21 has a unique format (full pesos, different labels)
    if period.upper() == "1Q21":
        return parse_balance_sheet_1q21(rows, period)

    # 3Q25+ compact format (3 columns)
    if ncols <= 4:
        return parse_balance_sheet_3q25_plus(rows, period)

    # Everything else
    return parse_balance_sheet_pre3q25(rows, period)


def parse_cash_flow(csv_path, period):
    """Parse cash flow statements.

    These have a dual-column layout: left side = operating activities,
    right side = investing/financing. Labels and values interleave.
    """
    rows = read_csv(csv_path)
    if not rows:
        return []

    # Reject non-cash-flow files (e.g. 3Q24 has a "Transaction summary")
    first_text = " ".join(rows[0]).lower() if rows else ""
    if "transaction" in first_text or "acquisition" in first_text[:50]:
        return []

    # Find the header/concept row (if present)
    concept_row_idx = None
    for i, row in enumerate(rows):
        joined = " ".join(row).lower()
        if "concept" in joined:
            concept_row_idx = i
            break

    # 2Q25-style: no concept row, 4-column layout (label, val, label, val)
    if concept_row_idx is None:
        return _parse_cash_flow_headerless(rows)

    pairs = _parse_cash_flow_standard(rows, concept_row_idx)
    return _maybe_rescale_to_thousands(pairs)


def _maybe_rescale_to_thousands(pairs):
    """Detect full-peso cash flows (2Q22) and convert to thousands.

    If the median absolute value of numeric entries exceeds 10M, the data
    is in full pesos and needs dividing by 1000.
    """
    numeric_vals = [abs(v) for _, v in pairs if isinstance(v, (int, float)) and v != 0]
    if not numeric_vals:
        return pairs
    numeric_vals.sort()
    median = numeric_vals[len(numeric_vals) // 2]
    if median > 10_000_000:
        return [(label, round(v / 1000, 0) if isinstance(v, (int, float)) else v) for label, v in pairs]
    return pairs


def _parse_cash_flow_headerless(rows):
    """Parse 4-column cash flow: col 0=left label, 1=left val, 2=right label, 3=right val."""
    results = []
    for row in rows:
        if len(row) < 2:
            continue
        if not any(c.strip() for c in row):
            continue

        # Skip footnotes
        if len(row[0]) > 80:
            continue

        left_label = row[0].strip()
        left_val_str = row[1].strip() if len(row) > 1 else ""
        right_label = row[2].strip() if len(row) > 2 else ""
        right_val_str = row[3].strip() if len(row) > 3 else ""

        if left_label:
            lv = clean_value(left_val_str)
            if lv is not None:
                results.append((left_label, lv))
            elif left_label.endswith(":"):
                results.append((left_label, None))

        if right_label:
            rv = clean_value(right_val_str)
            if rv is not None:
                results.append((right_label, rv))
            elif right_label.endswith(":"):
                results.append((right_label, None))

    return results


def _parse_cash_flow_standard(rows, concept_row_idx):
    """Parse cash flow with a 'Concept' header row. Values are embedded in
    cells alongside right-side labels."""
    results = []
    for row in rows[concept_row_idx + 1:]:
        if len(row) < 2:
            continue
        if not any(c.strip() for c in row):
            continue

        left_label = row[0].strip()

        # Skip noise: fragment labels (single words like "development.", "instruments")
        # and narrative text
        if left_label and not left_label.endswith(":"):
            if len(left_label) < 4 or (len(left_label) > 80):
                continue
            if left_label.lower() in ("concept", "instruments", "development.",
                                       "under development.", "acquisitions"):
                continue

        left_val = None
        left_val_col = None
        for ci in range(1, min(3, len(row))):
            cell = row[ci].strip()
            num_match = re.match(r"^([\$\(\)\-\d,.]+)\s*(.*)", cell)
            if num_match:
                cv = clean_value(num_match.group(1))
                if cv is not None:
                    left_val = cv
                    left_val_col = ci
                    break
            if cell.startswith("(") and ")" in cell:
                paren_end = cell.index(")") + 1
                cv = clean_value(cell[:paren_end])
                if cv is not None:
                    left_val = cv
                    left_val_col = ci
                    break

        if left_label and left_val is not None:
            results.append((left_label, left_val))
        elif left_label and left_label.endswith(":"):
            results.append((left_label, None))

        # --- Right side ---
        right_label = None
        right_val = None

        for ci in range(1, len(row)):
            cell = row[ci].strip()
            if not cell:
                continue

            if left_val_col is not None and ci == left_val_col:
                num_match = re.match(r"^[\$\(\)\-\d,.]+\s+(.*)", cell)
                if num_match:
                    trailing = num_match.group(1).strip()
                    if trailing and len(trailing) > 2:
                        right_label = trailing
                continue

            if ci > (left_val_col or 1):
                cv = clean_value(cell)
                if cv is not None and right_label:
                    right_val = cv
                elif cv is not None and not right_label:
                    right_val = cv
                elif cell and not right_label:
                    right_label = cell
                elif cell and right_label and right_val is None:
                    right_label = right_label + " " + cell

        if right_label and right_val is not None:
            results.append((right_label, right_val))
        elif right_label and right_label.endswith(":"):
            results.append((right_label, None))

    return results


# ---------------------------------------------------------------------------
# Label alignment
# ---------------------------------------------------------------------------

def align_labels(all_data):
    """Merge label lists across quarters into a single ordered master list.

    all_data: [(period, [(label, value), ...]), ...]

    Returns: (master_labels, {period: {label: value}})
    """
    master = []
    master_norm = []
    period_maps = {}

    for period, pairs in all_data:
        pmap = {}
        prev_master_idx = -1

        for label, val in pairs:
            nkey = normalize_label(label)

            # Find match in master list
            found_idx = None
            for mi, mn in enumerate(master_norm):
                if mn == nkey:
                    found_idx = mi
                    break

            if found_idx is not None:
                # Use the canonical label from master
                canonical = master[found_idx]
                pmap[canonical] = val
                prev_master_idx = found_idx
            else:
                # New label — insert after the last matched position
                insert_pos = prev_master_idx + 1
                master.insert(insert_pos, label)
                master_norm.insert(insert_pos, nkey)
                pmap[label] = val
                prev_master_idx = insert_pos

        period_maps[period] = pmap

    return master, period_maps


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------

HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_FONT_WHITE = Font(bold=True, size=11, color="FFFFFF")
SECTION_FONT = Font(bold=True, size=10)
NUMBER_FMT = '#,##0'
NEG_NUMBER_FMT = '#,##0;(#,##0)'


def write_model_tab(wb, tab_name, master_labels, period_maps, periods):
    ws = wb.create_sheet(title=tab_name[:31])

    # Row 1: header — col A = blank or "Thousand Pesos", cols B+ = periods
    ws.cell(row=1, column=1, value="Thousand Pesos").font = HEADER_FONT_WHITE
    ws.cell(row=1, column=1).fill = HEADER_FILL
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="left")

    for pi, period in enumerate(periods):
        cell = ws.cell(row=1, column=pi + 2, value=period)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for ri, label in enumerate(master_labels):
        row_num = ri + 2
        label_cell = ws.cell(row=row_num, column=1, value=label)

        # Bold section headers (labels ending with ":")
        is_section = label.rstrip().endswith(":") or (
            clean_value(label) is None and
            all(period_maps.get(p, {}).get(label) is None for p in periods)
        )
        if is_section:
            label_cell.font = SECTION_FONT

        for pi, period in enumerate(periods):
            val = period_maps.get(period, {}).get(label)
            if val is None:
                continue
            cell = ws.cell(row=row_num, column=pi + 2)
            if isinstance(val, str) and val.endswith("%"):
                cell.value = val
                cell.alignment = Alignment(horizontal="right")
            elif isinstance(val, (int, float)):
                cell.value = val
                cell.number_format = NEG_NUMBER_FMT
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.value = val

    # Freeze pane at B2 (scroll data, keep labels and header visible)
    ws.freeze_panes = "B2"

    # Auto-width
    # Column A (labels): measure content
    max_a = 20
    for ri in range(1, len(master_labels) + 2):
        cell = ws.cell(row=ri, column=1)
        if cell.value:
            max_a = max(max_a, len(str(cell.value)))
    ws.column_dimensions["A"].width = min(max_a + 2, 60)

    # Period columns: fixed width
    for pi in range(len(periods)):
        ws.column_dimensions[get_column_letter(pi + 2)].width = 14

    return ws


# ---------------------------------------------------------------------------
# Orchestration
# ---------------------------------------------------------------------------

def discover_csvs(parent_dir, stmt_type):
    """Find all primary CSVs for a statement type, return [(period, path)] sorted."""
    parent = Path(parent_dir)
    results = []
    skip = SKIP_PERIODS.get(stmt_type, set())

    for tables_dir in parent.glob("*_tables"):
        if not tables_dir.is_dir():
            continue
        m = re.match(r"(\d[Qq]\d{2})", tables_dir.name)
        if not m:
            continue
        period = m.group(1).upper()
        if period in skip:
            continue

        csv_file = tables_dir / f"{stmt_type}_{period.lower()}.csv"
        if not csv_file.exists():
            csv_file = tables_dir / f"{stmt_type}_{period}.csv"
        if not csv_file.exists():
            continue

        results.append((period, csv_file))

    results.sort(key=lambda x: period_sort_key(x[0]))
    return results


PARSERS = {
    "income_statement": parse_income_statement,
    "balance_sheet": parse_balance_sheet,
    "cash_flow": parse_cash_flow,
    "noi_ebitda": parse_noi_ebitda,
    "ffo_affo": parse_ffo_affo,
}


def build_workbook(parent_dir, output_path=None):
    parent = Path(parent_dir)
    if output_path is None:
        output_path = parent / "fibrasoma_model.xlsx"
    else:
        output_path = Path(output_path)

    wb = Workbook()
    wb.remove(wb.active)

    for stmt_type, tab_name in TAB_CONFIG:
        entries = discover_csvs(parent_dir, stmt_type)
        if not entries:
            print(f"  {tab_name}: no CSVs found, skipping", file=sys.stderr)
            continue

        parser = PARSERS[stmt_type]
        periods = [e[0] for e in entries]
        print(f"  {tab_name}: {len(entries)} quarters ({periods[0]}–{periods[-1]})", file=sys.stderr)

        all_data = []
        for period, csv_path in entries:
            pairs = parser(csv_path, period)
            if not pairs:
                print(f"    WARNING: no data extracted from {csv_path.name}", file=sys.stderr)
                continue
            all_data.append((period, pairs))

        if not all_data:
            continue

        actual_periods = [p for p, _ in all_data]
        master_labels, period_maps = align_labels(all_data)
        write_model_tab(wb, tab_name, master_labels, period_maps, actual_periods)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    print(f"\nWorkbook saved: {output_path}", file=sys.stderr)
    print(f"Tabs: {[ws.title for ws in wb.worksheets]}", file=sys.stderr)
    return str(output_path)


def main():
    import argparse
    parser = argparse.ArgumentParser(description="Build financial model workbook from extracted tables")
    parser.add_argument("tables_parent_dir", help="Directory containing *_tables/ subdirectories")
    parser.add_argument("--output", help="Output Excel file path")
    args = parser.parse_args()

    result = build_workbook(args.tables_parent_dir, args.output)
    print(result)


if __name__ == "__main__":
    main()
